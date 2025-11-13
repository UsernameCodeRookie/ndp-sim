#!/usr/bin/env python3
"""生成匹配 gemm.json 配置的寄存器映射表 Excel"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from dataclasses import dataclass
from typing import List, Optional, Literal, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@dataclass
class BitField:
    name: str
    width: int
    repeat: int = 1
    meaning: str = ""
    port: str = ""
    remark: str = ""
    align_lsb_to: Optional[int] = None
    fixed_msb: Optional[int] = None

def layout_fields(
    fields: List[BitField],
    total_bits: int,
    pack_direction: Literal["msb_to_lsb", "lsb_to_msb"] = "msb_to_lsb",
    start_bit: Optional[int] = None,
) -> List[Tuple[str, str, str, str, str]]:
    if pack_direction == "msb_to_lsb":
        cur = (total_bits - 1) if start_bit is None else start_bit
        step = -1
    else:
        cur = 0 if start_bit is None else start_bit
        step = +1

    rows = []

    def take_range(width: int) -> Tuple[int, int]:
        nonlocal cur
        if pack_direction == "msb_to_lsb":
            msb = cur
            lsb = cur - width + 1
            if lsb < 0:
                raise ValueError(f"位宽溢出：{msb}:{lsb} 超出 0..{total_bits-1}")
            cur = lsb - 1
        else:
            lsb = cur
            msb = cur + width - 1
            if msb >= total_bits:
                raise ValueError(f"位宽溢出：{msb}:{lsb} 超出 0..{total_bits-1}")
            cur = msb + 1
        return msb, lsb

    for f in fields:
        segs = []
        for _ in range(f.repeat):
            msb, lsb = take_range(f.width)
            segs.append((msb, lsb))

        if f.repeat > 1:
            msb_all = segs[0][0]
            lsb_all = segs[-1][1]
            field_bits = f"{f.repeat}×{f.width}bit[{msb_all}:{lsb_all}]"
        else:
            field_bits = f"{f.width}bit[{segs[0][0]}:{segs[0][1]}]"

        rows.append((field_bits, f.name, f.meaning, f.remark, f.port))

    return rows

def write_section(ws, start_row, section_title, rows):
    start = start_row
    row = start_row
    for field_bits, name, meaning, remark, port in rows:
        ws.cell(row=row, column=3, value=field_bits)
        ws.cell(row=row, column=4, value=name)
        ws.cell(row=row, column=5, value=port)
        ws.cell(row=row, column=6, value=meaning)
        ws.cell(row=row, column=7, value=remark)
        row += 1

    if row > start:
        ws.merge_cells(start_row=start, start_column=2, end_row=row-1, end_column=2)
        ws.cell(row=start, column=2, value=section_title)
        ws.cell(row=start, column=2).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.cell(row=start, column=2).font = Font(bold=True)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start, row):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    return row

def build_register_map_excel(path, groups_spec):
    wb = Workbook()
    ws = wb.active
    ws.title = "Register Map"

    headers = ["大组名称", "模块名称", "字段", "配置名", "硬件端口名", "备注", "默认值"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)
        ws.cell(row=1, column=i).alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="999999")
        ws.cell(row=1, column=i).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "C2"

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 22
    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(4)].width = 54
    ws.column_dimensions[get_column_letter(5)].width = 30
    ws.column_dimensions[get_column_letter(6)].width = 50
    ws.column_dimensions[get_column_letter(7)].width = 10

    head_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = head_fill

    current_row = 2
    for group in groups_spec:
        group_title = group["group_title"]
        sections = group["sections"]
        
        group_total_rows = 0
        group_rows_list = []
        for sec in sections:
            rows = layout_fields(
                sec["fields"],
                total_bits=sec.get("total_bits", 128),
                pack_direction=sec.get("pack_direction", "lsb_to_msb"),
            )
            group_rows_list.append(rows)
            group_total_rows += len(rows)
        
        if group_total_rows > 0:
            start_row = current_row
            end_row = current_row + group_total_rows - 1
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1, value=group_title)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        
        for i, sec in enumerate(sections):
            current_row = write_section(ws, current_row, sec["title"], group_rows_list[i])

    wb.save(path)

# 按大组组织sections - 针对 gemm.json 配置
groups = [
    {
        "group_title": "CONFIG",
        "sections": [
            {
                "title": "CONFIG",
                "total_bits": 8,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("config", width=8, port="config", meaning="iga、se、buffer、special等模块使能配置"),
                ],
            }
        ]
    },
    {
        "group_title": "Index Generation",
        "sections": [
            {
                "title": "8*DRAM LC",
                "total_bits": 48, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("dram_loop_configs.src_id",            width=2,          port="iga_lc_src_id",           meaning="LC0-LC7的src_id，null时置为00"),
                    BitField("dram_loop_configs.outmost_loop",      width=1,          port="iga_lc_outmost_loop",     meaning="src_id=null时outmost_loop可设为1"),
                    BitField("dram_loop_configs.start",             width=13,         port="iga_lc_initial_value",    meaning="循环起始值"),
                    BitField("dram_loop_configs.stride",            width=13,         port="iga_lc_stride_value",     meaning="循环步进值"),
                    BitField("dram_loop_configs.end",               width=13,         port="iga_lc_end_value",        meaning="循环结束值"),
                    BitField("dram_loop_configs.last_index",        width=3,          port="iga_lc_index",            meaning="last_index端口索引"),
                ],
            },
            {
                "title": "4*BUFFER ROW LC",
                "total_bits": 40,  
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("buffer_loop_configs.ROW_LC.src_id",       width=3,          port="iga_row_lc_src_id",           meaning="ROW_LC的输入可以是DRAM LC"),
                    BitField("buffer_loop_configs.ROW_LC.start",        width=3,          port="iga_row_lc_initial_value",    meaning="循环起始值"),
                    BitField("buffer_loop_configs.ROW_LC.stride",       width=3,          port="iga_row_lc_stride_value",     meaning="循环步进值"),
                    BitField("buffer_loop_configs.ROW_LC.end",          width=3,          port="iga_row_lc_end_value",        meaning="循环结束值"),
                    BitField("buffer_loop_configs.ROW_LC.last_index",   width=3,          port="iga_row_lc_index" ,           meaning="last_index端口索引"),
                ],
            },
            {
                "title": "4*BUFFER COL LC",
                "total_bits": 40,         
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("buffer_loop_configs.COL_LC.src_id",       width=3,          port="iga_col_lc_src_id",           meaning="COL_LC的输入可以是DRAM LC"),
                    BitField("buffer_loop_configs.COL_LC.start",        width=6,          port="iga_col_lc_initial_value",    meaning="循环起始值"),
                    BitField("buffer_loop_configs.COL_LC.stride",       width=6,          port="iga_col_lc_stride_value",     meaning="循环步进值"),
                    BitField("buffer_loop_configs.COL_LC.end",          width=6,          port="iga_col_lc_end_value",        meaning="循环结束值"),
                    BitField("buffer_loop_configs.COL_LC.last_index",   width=3,          port="iga_col_lc_index" ,           meaning="last_index端口索引"),
                ],
            },
            {
                "title": "8*LC PE",
                "total_bits": 104,  
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("lc_pe_configs.inport.mode",          width=2,            port="iga_pe_inport_mode",          meaning="null:00, buffer:01, keep:10, constant:11"),
                    BitField("lc_pe_configs.inport.keep_last_index", width=3,          port="iga_pe_keep_last_index",      meaning="各inport的last_index"),
                    BitField("lc_pe_configs.inport.src_id",        width=3,            port="iga_pe_src_id",               meaning="LC_PE的输入源ID"),
                    BitField("lc_pe_configs.alu_opcode",           width=2,            port="iga_pe_alu_opcode",           meaning="add:0, mul:1, max:2"),
                    BitField("lc_pe_configs.inport.cfg_constant_pos",width=12,         port="iga_pe_constant_value",       meaning="3个inport的常量值，每个12bit"),
                ],
            },
        ]
    },
    {
        "group_title": "LSU",
        "sections": [
            {
                "title": "1*Read Memory Stream Engine (stream0)",
                "total_bits": 720,            
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("",                                                            width=4,                    port="padding",                     meaning="填充位对齐"),
                    BitField("stream_engine.stream.memory_AG.idx_keep_mode",                width=6,                    port="mse_mem_idx_mode",            meaning="idx0,idx1,idx2的模式：buffer(1)、keep(2)、constant(3)"),
                    BitField("stream_engine.stream.memory_AG.idx_keep_last_index",          width=9,                    port="mse_mem_idx_keep_last_index", meaning="3个idx的last_index，每个3bit"),
                    BitField("stream_engine.stream.memory_AG.idx",                          width=12,                   port="mem_inport_src_encode",       meaning="idx0,idx1,idx2来源DRAM LC编号，每个4bit"),
                    BitField("stream_engine.stream.memory_AG.idx_constant",                 width=24,                   port="mse_mem_idx_constant",        meaning="3个idx的常量值，每个8bit"),
                    BitField("stream_engine.stream.buffer_AG.idx_keep_mode",                width=2,                    port="mse_buf_idx_keep_mode",       meaning="buffer AG 2个port的模式"),
                    BitField("stream_engine.stream.buffer_AG.idx_keep_last_index",          width=6,                    port="mse_buf_idx_keep_last_index", meaning="buffer AG 2个port的last_index"),
                    BitField("stream_engine.stream.stream.ping_pong",                       width=1,                    port="mse_pingpong_enable",         meaning="ping-pong使能"),
                    BitField("stream_engine.stream.stream.pingpong_last_index",             width=3,                    port="mse_pingpong_last_index",     meaning="pingpong切换的last_index"),
                    BitField("stream_engine.stream.memory_AG.base_addr",                    width=29,                   port="mse_stream_base_addr",        meaning="内存基地址"),
                    BitField("stream_engine.stream.memory_AG.idx_size",                     width=24,                   port="mse_transaction_layout_size", meaning="idx0_size、idx1_size、idx2_size，每个8bit"),
                    BitField("",                                                            width=9,                    port="mse_transaction_layout_size_log", meaning="计算: log2(idx2_size)和log2(idx1_size*idx2_size)"),
                    BitField("",                                                            width=8,                    port="mse_transaction_total_size",  meaning="计算: idx0_size*idx1_size*idx2_size"),
                    BitField("stream_engine.stream.memory_AG.dim_stride",                   width=72,                   port="mse_transaction_mult",        meaning="idx0_stride、idx1_stride、idx2_stride，每维3个值共24bit"),
                    BitField("stream_engine.stream.memory_AG.address_remapping",            width=64,                   port="mse_map_matrix_b",            meaning="地址重映射表，16个4bit值"),
                    BitField("stream_engine.stream.memory_AG.padding_reg_value",            width=8,                    port="mse_padding_reg_value",       meaning="padding填充值"),
                    BitField("stream_engine.stream.memory_AG.padding_enable",               width=3,                    port="mse_padding_valid",           meaning="3个idx的padding使能"),
                    BitField("stream_engine.stream.memory_AG.idx_padding_range.low_bound",  width=42,                   port="mse_padding_low_bound",       meaning="3个idx的padding低位界，每个14bit"),
                    BitField("stream_engine.stream.memory_AG.idx_padding_range.up_bound",   width=42,                   port="mse_padding_up_bound",        meaning="3个idx的padding高位界，每个14bit"),
                    BitField("stream_engine.stream.memory_AG.tailing_enable",               width=3,                    port="mse_branch_valid",            meaning="3个idx的tailing使能"),
                    BitField("stream_engine.stream.memory_AG.idx_tailing_range.low",        width=42,                   port="mse_branch_low_bound",        meaning="3个idx的有效下边界，每个14bit"),
                    BitField("stream_engine.stream.memory_AG.idx_tailing_range.up",         width=42,                   port="mse_branch_up_bound",         meaning="3个idx的有效上边界，每个14bit"),
                    BitField("stream_engine.stream.buffer_AG.spatial_stride",               width=256,                  port="mse_buf_spatial_stride",      meaning="buffer ag展开的16路offset，每路16bit"),
                    BitField("stream_engine.stream.buffer_AG.spatial_size",                 width=5,                    port="mse_buf_spatial_size",        meaning="buffer ag需要展开的地址个数"),
                ],
            },
            {
                "title": "1*Neighbor Stream Engine",
                "total_bits": 36,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("stream_engine.n2n.mem_loop",         width=4,            port="nse_cnt_size",            meaning="内存循环大小"),
                    BitField("stream_engine.n2n.enable",           width=1,            port="nse_enable",              meaning="使能标志"),
                    BitField("",                                   width=2,            port="nse_stream_id",           meaning="流ID(未使用)"),
                    BitField("stream_engine.n2n.src_slice_sel",    width=1,            port="nse_in_src_slice_sel",    meaning="1表示跳4个slice，0表示不跳"),
                    BitField("stream_engine.n2n.dst_slice_sel",    width=1,            port="nse_out_dst_slice_sel",   meaning="目标slice选择"),
                    BitField("",                                   width=3,            port="nse_src_buf_ping_idx",    meaning="源buffer ping索引"),
                    BitField("",                                   width=3,            port="nse_src_buf_pong_idx",    meaning="源buffer pong索引"),
                    BitField("",                                   width=3,            port="nse_dst_buf_ping_idx",    meaning="目标buffer ping索引"),
                    BitField("",                                   width=3,            port="nse_dst_buf_pong_idx",    meaning="目标buffer pong索引"),
                ],
            },
        ]
    },
    {
        "group_title": "Buffer",
        "sections": [
            {
                "title": "6*Buffer_Manager_Cluster",
                "total_bits": 28,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("buffer_config.buffer.dst_port",           width=1,                            port="buf_wr_src_id",           meaning="0: SpecArray, 1: GeneArray"),
                    BitField("buffer_config.buffer.buffer_life_time",   width=2,                            port="buffer_life_time",        meaning="buffer被复用多少次"),
                    BitField("buffer_config.buffer.mode",               width=1,                            port="buffer_mode",             meaning="mode0:for(life)for(row) mode1:for(row)for(life)"),
                    BitField("buffer_config.buffer.mask",               width=8,                            port="buffer_mask",             meaning="8个bank的使能掩码"),
                ],
            }
        ]
    },
    {
        "group_title": "SA",
        "sections": [
            {
                "title": "1*PE Config",
                "total_bits": 14, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("special_array.data_type",         width=2,            port="sa_pe_computation_data_type", meaning="fp16:1 int8:0"),
                    BitField("special_array.index_end",         width=3,            port="sa_pe_config_last_index",     meaning="PE的输出什么时候有效"),
                    BitField("special_array.bias_enable",       width=1,            port="sa_pe_bias_enable",           meaning="bias使能"),
                ],
            },
            {
                "title": "3*Inport Config",
                "total_bits": 15, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("special_array.inport.enable",                 width=1,                        port="sa_inport_enable",                meaning="inport使能"),
                    BitField("special_array.inport.pingpong_en",            width=1,                        port="sa_inport_pingpong_en",           meaning="pingpong使能"),
                    BitField("special_array.inport.pingpong_last_index",    width=3,                        port="sa_inport_pingpong_last_index",   meaning="pingpong切换的last_index"),
                ],
            },
            {
                "title": "1*Outport Config",
                "total_bits": 2,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("special_array.outport.mode",          width=1,            port="sa_outport_major",            meaning="col：0  row:1"),
                    BitField("special_array.outport.fp32to16",      width=1,            port="sa_outport_fp32to16",         meaning="true: 1 false: 0"),
                ],
            },
        ]
    },
]

if __name__ == "__main__":
    output_file = "register_map_gemm.xlsx"
    build_register_map_excel(output_file, groups)
    print(f"生成完成：{output_file}")
    print(f"此Excel对照表匹配配置文件：data/gemm.json")
