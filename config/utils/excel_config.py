from dataclasses import dataclass
from typing import List, Optional, Literal, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math
from config.utils.config_parameters import *

# ===================== 1) 位段描述与自动排布 =====================

@dataclass
class BitField:
    # 逻辑名称（用于“含义”列左侧的小字段名）
    name: str
    # 每个子段的位宽
    width: int
    # 重复次数（例如 16×5bit 的“16”）
    repeat: int = 1
    # 备注/含义文本
    meaning: str = ""
    port: str = ""
    remark: str = ""
    # 可选：对齐到某个 lsb（例如想让该字段落在偶数边界）
    align_lsb_to: Optional[int] = None
    # 可选：固定起始 msb（高位开始）；若提供则将从此处开始分配该字段（常用于跨寄存器手动锚定）
    fixed_msb: Optional[int] = None

def _align_down(lsb: int, align_to: int) -> int:
    """把 lsb 向下取整到 align_to 的倍数边界（返回新的 lsb）。"""
    if align_to is None or align_to <= 1:
        return lsb
    return (lsb // align_to) * align_to

def layout_fields(
    fields: List[BitField],
    total_bits: int,
    pack_direction: Literal["msb_to_lsb", "lsb_to_msb"] = "msb_to_lsb",
    start_bit: Optional[int] = None,
) -> List[Tuple[str, str, str, str, str]]:
    """
    根据位宽/重复次数自动布置位段，返回用于 Excel 的行：
    (field_bits_str, name, meaning, remark)

    - total_bits: 本寄存器（或打包域）的总位宽（例如 135、128 等）
    - pack_direction: 缺省从 MSB 向 LSB 递减（硬件常见做法）
    - start_bit: 起始 bit（默认 msb_to_lsb 模式为 total_bits-1；lsb_to_msb 为 0）
    """
    if pack_direction == "msb_to_lsb":
        cur = (total_bits - 1) if start_bit is None else start_bit
        step = -1
    else:
        cur = 0 if start_bit is None else start_bit
        step = +1

    rows = []

    def take_range(width: int) -> Tuple[int, int]:
        nonlocal cur
        if pack_direction == "msb_to_lsb":
            msb = cur
            lsb = cur - width + 1
            if lsb < 0:
                raise ValueError(f"位宽溢出：{msb}:{lsb} 超出 0..{total_bits-1}")
            cur = lsb - 1
        else:
            lsb = cur
            msb = cur + width - 1
            if msb >= total_bits:
                raise ValueError(f"位宽溢出：{msb}:{lsb} 超出 0..{total_bits-1}")
            cur = msb + 1
        return msb, lsb

    for f in fields:
        # 处理对齐
        if f.align_lsb_to and pack_direction == "msb_to_lsb":
            # 先预留 f.width，以便对齐 lsb
            future_lsb = cur - (f.width - 1)
            aligned_lsb = _align_down(future_lsb, f.align_lsb_to)
            # 如果对齐导致移动，更新 cur
            cur = aligned_lsb + (f.width - 1)
        elif f.fixed_msb is not None:
            # 强制锚定（仅支持 msb_to_lsb）
            cur = f.fixed_msb

        # 逐个重复子段分配
        segs = []
        for _ in range(f.repeat):
            msb, lsb = take_range(f.width)
            segs.append((msb, lsb))

        # 把重复段合并成 "16×5bit[79:0]" 风格的展示文本
        if f.repeat > 1:
            # 连续重复的段通常是相邻拼接的，整体范围用首末覆盖
            msb_all = segs[0][0]
            lsb_all = segs[-1][1]
            field_bits = f"{f.repeat}×{f.width}bit[{msb_all}:{lsb_all}]"
        else:
            field_bits = f"{f.width}bit[{segs[0][0]}:{segs[0][1]}]"

        rows.append((field_bits, f.name, f.meaning, f.remark, f.port))

    return rows

# ===================== 2) 写 Excel（保持你原来的“图一”版式） =====================

def write_section(ws, start_row, section_title, rows):
    start = start_row
    row = start_row
    for field_bits, name, meaning, remark, port in rows:
        ws.cell(row=row, column=3, value=field_bits)  # 字段
        ws.cell(row=row, column=4, value=name)        # 含义-字段名
        ws.cell(row=row, column=5, value=port)
        ws.cell(row=row, column=6, value=meaning)     # 含义-说明
        ws.cell(row=row, column=7, value=remark)      # 备注
        row += 1

    # 左侧模块名合并
    if row > start:
        ws.merge_cells(start_row=start, start_column=2, end_row=row-1, end_column=2)
        ws.cell(row=start, column=2, value=section_title)
        ws.cell(row=start, column=2).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.cell(row=start, column=2).font = Font(bold=True)

    # 样式
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start, row):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    return row

def build_register_map_excel(path, groups_spec):
    """
    groups_spec: [
        {
            "group_title": "引擎配置",  # 最左列的大组标题
            "sections": [  # 该组包含的sections
                {
                    "title": "Memory Stream Engine",
                    "total_bits": 1024,
                    "fields": [BitField(...)]
                },
                ...
            ]
        },
        ...
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Register Map"

    # 表头（新增第一列为大组标题）
    headers = ["大组名称", "模块名称", "字段", "配置名", "硬件端口名", "备注", "默认值"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)
        ws.cell(row=1, column=i).alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="999999")
        ws.cell(row=1, column=i).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "C2"

    # 列宽设置
    ws.column_dimensions[get_column_letter(1)].width = 18  # 大组名称列
    ws.column_dimensions[get_column_letter(2)].width = 22  # 模块名称列
    ws.column_dimensions[get_column_letter(3)].width = 20  # 字段列
    ws.column_dimensions[get_column_letter(4)].width = 54  # 配置名列
    ws.column_dimensions[get_column_letter(5)].width = 30  # 硬件端口名列
    ws.column_dimensions[get_column_letter(6)].width = 50  # 备注列
    ws.column_dimensions[get_column_letter(7)].width = 10  # 默认值列

    head_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = head_fill

    r = 2
    current_row = 2
    # 遍历每个大组
    for group in groups_spec:
        group_title = group["group_title"]
        sections = group["sections"]
        
        # 计算该大组包含的总行数
        group_total_rows = 0
        group_rows_list = []
        for sec in sections:
            rows = layout_fields(
                sec["fields"],
                total_bits=sec.get("total_bits", 128),
                pack_direction=sec.get("pack_direction", "lsb_to_msb"),
            )
            group_rows_list.append(rows)
            group_total_rows += len(rows)
        
        # 合并大组标题（第一列）
        if group_total_rows > 0:
            start_row = current_row
            end_row = current_row + group_total_rows - 1
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(row=start_row, column=1, value=group_title)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        
        # 写入该大组下的所有sections
        for i, sec in enumerate(sections):
            current_row = write_section(ws, current_row, sec["title"], group_rows_list[i])

    wb.save(path)

# 按大组组织sections
groups = [
    {
        "group_title": "CONFIG",
        "sections": [
            {
                "title": "CONFIG",
                "total_bits": 4,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("config", width=4, port="", meaning="iga、lsu、sa、ga"),
                ],
            }
        ]
    },
    {
        "group_title": "Index Generation",
        "sections": [
            {
                "title": "8*DRAM LC",
                "total_bits": 1024, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("",                                    width=1,                            port="iga_lc_src_enable",       meaning="DRAM LC的输入时DRAM LC，配置字src_id=null时iga_lc_src_enable置为0",remark="0"),
                    BitField("dram_loop_configs.src_id",            width=IGA_LC_SRC_ID_WIDTH,          port="iga_lc_src_id",           meaning="需要根据当前配置的LC的id推测src_id转换成数字是多少",remark="0"),
                    BitField("dram_loop_configs.outmost_loop",      width=IGA_LC_OUTMOST_LOOP,          port="iga_lc_outmost_loop",     meaning="inport=null时outmost_loop设为1",remark=""),
                    BitField("dram_loop_configs.start",             width=IGA_LC_INITIAL_VALUE_WIDTH,   port="iga_lc_initial_value",    remark=""),
                    BitField("dram_loop_configs.end",               width=IGA_LC_STRIDE_VALUE_WIDTH,    port="iga_lc_stride_value",     remark=""),
                    BitField("dram_loop_configs.stride",            width=IGA_LC_END_VALUE_WIDTH,       port="iga_lc_end_value",        remark=""),
                    BitField("dram_loop_configs.last_index",        width=IGA_LC_CFG_PORT_WIDTH,        port="iga_lc_index",            remark=""),
                ],
            },
            {
                "title": "4*BUFFER ROW LC",
                "total_bits": 1024,  
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("buffer_loop_configs.ROW_LC.src_id",       width=IGA_ROW_LC_SRC_ID_WIDTH,          port="iga_row_lc_src_id",           meaning="ROW_LC的输入可以是COL_LC也可以是DRAM LC",remark="0"),
                    BitField("buffer_loop_configs.ROW_LC.start",        width=IGA_ROW_LC_INITIAL_VALUE_WIDTH,   port="iga_row_lc_initial_value",    meaning="",remark=""),
                    BitField("buffer_loop_configs.ROW_LC.stride",       width=IGA_ROW_LC_STRIDE_VALUE_WIDTH,    port="iga_row_lc_stride_value",     meaning="",remark=""),
                    BitField("buffer_loop_configs.ROW_LC.end",          width=IGA_ROW_LC_END_VALUE_WIDTH,       port="iga_row_lc_end_value",        meaning="",remark=""),
                    BitField("buffer_loop_configs.ROW_LC.last_index",   width=IGA_ROW_LC_CFG_PORT_WIDTH,        port="iga_row_lc_index" ,           meaning="",remark=""),
                ],
            },
            {
                "title": "4*BUFFER COL LC",
                "total_bits": 1024,         
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("buffer_loop_configs.COL_LC.src_id",       width=IGA_COL_LC_SRC_ID_WIDTH,          port="iga_col_lc_src_id",           meaning="COL_LC的输入可以是ROW_LC也可以是DRAM LC",remark="0"),
                    BitField("buffer_loop_configs.COL_LC.start",        width=IGA_COL_LC_INITIAL_VALUE_WIDTH,   port="iga_col_lc_initial_value",    meaning="",remark=""),
                    BitField("buffer_loop_configs.COL_LC.stride",       width=IGA_COL_LC_STRIDE_VALUE_WIDTH,    port="iga_col_lc_stride_value",     meaning="",remark=""),
                    BitField("buffer_loop_configs.COL_LC.end",          width=IGA_COL_LC_END_VALUE_WIDTH,       port="iga_col_lc_end_value",        meaning="",remark=""),
                    BitField("buffer_loop_configs.COL_LC.last_index",   width=IGA_COL_LC_CFG_PORT_WIDTH,        port="iga_col_lc_index" ,           meaning="",remark=""),
                ],
            },
            {
                "title": "4*LC PE",
                "total_bits": 1024,  
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("LC_PE.inport",            width=IGA_PE_INPORT_NUM*IGA_PE_SRC_ID_WIDTH,        port="iga_pe_src_id",               meaning="LC_PE的输入可以是DRAM LC也可以是LC_PE",remark=""),
                    BitField("LC_PE.inport_mode",       width=IGA_PE_INPORT_NUM*IGA_PE_INPORT_MODE_WIDTH,   port="iga_pe_inport_mode",          meaning="null: 2'00, buffer: 2'01, keep: 2'10, constant: 2'11"),
                    BitField("LC_PE.inport_last_index", width=IGA_PE_INPORT_NUM*PORT_LAST_INDEX,            port="iga_pe_keep_last_index",      remark=""),
                    BitField("LC_PE.opcode",            width=IGA_PE_ALU_OPCODE_WIDTH,                      port="iga_pe_alu_opcode",           meaning="add: 0, mul: 1, max: 2", remark=""),
                    BitField("LC_PE.inport_enable",     width=IGA_PE_INPORT_NUM,                            port="iga_pe_inport_enable",        remark=""),
                    BitField("LC_PE.constant_value",    width=IGA_PE_INPORT_NUM*IGA_PE_CONSTANT_VALUE_WIDTH,port="iga_pe_constant_value",       remark=""),
                    BitField("LC_PE.constant_valid",    width=IGA_PE_INPORT_NUM,                            port="iga_pe_constant_valid",       remark=""),
                ],
            },
        ]
    },
    {
        "group_title": "LSU",
        "sections": [
            {
                "title": "3*Read Memory Stream Engine",
                "total_bits": 1024,            
                "pack_direction": "lsb_to_msb",
                "fields": [
                    # 配置点
                    BitField("",                                                            width=1,                    port="mse_enable",           meaning="配置了就为1没配就为0",               remark=""),
                    BitField("stream_engine.stream.memory_AG.mode",                         width=1,                    meaning="read: 需要配padding和tailing  write: 不需要配",               remark=""),
                    BitField("stream_engine.stream.memory_AG.base_addr",                    width=ADDR_WIDTH,           port="mse_stream_base_addr",        meaning="基地址",                      remark="" ),
                    BitField("stream_engine.stream.memory_AG.idx_size",                     width=IDX_SIZE_WIDTH,       port="mse_transaciton_layout_size",meaning="idx0_size、idx1_size、idx2_size 真实值减1"  ),
                    BitField("stream_engine.stream.memory_AG.dim_stride",                   width=STRIDE_WIDTH,         port="mse_transaciton_mult",        meaning="idx0_stride、idx1_stride、idx2_stride"    ),
                    BitField("stream_engine.stream.memory_AG.padding_enable",               width=AG_PORT_NUM,          port="mse_padding_valid",          meaning="idx_padding使能,置为1的话使能的输入端口padding"         ,remark=""),
                    BitField("stream_engine.stream.memory_AG.padding_reg_value",            width=MSE_PADDING_VALUE_WIDTH, port="mse_padding_reg_value",    meaning="padding填充值"         ,remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_padding_range low_bound",  width=PAD_RANGE_WIDTH,      port="mse_padding_low_bound",       meaning="idx0、idx1、idex2 padding低位界"     ,remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_padding_range up_bound",   width=PAD_RANGE_WIDTH,      port="mse_padding_up_bound",        meaning="idx0、idx1、idex2 padding高位界"     ,remark=""),
                    BitField("stream_engine.stream.memory_AG.tailing_enable",               width=AG_PORT_NUM,          port="mse_branch_valid",          meaning="idx_tailing使能,置为1的话使能的输入端口tailing"         ,remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_tailing_range low",        width=TAIL_RANGE_WIDTH,     port="mse_branch_low_bound",        meaning="idx0、idx1、idex2 有效上边界"         ,remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_tailing_range up",         width=TAIL_RANGE_WIDTH,     port="mse_branch_up_bound",         meaning="idx0、idx1、idex2 有效下边界"         ,remark=""),
                    BitField("stream_engine.stream.memory_AG.address_remapping",            width=REMAP_WIDTH,          port="mse_map_matrix_b",            meaning="地址重映射表/掩码"     ,remark="[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]" ),
                    
                    BitField("stream_engine.stream.stream.ping_pong",                       width=PING_PONG_WIDTH,      port="mse_pingpong_enable",         meaning="ping-pong使能", remark=""),
                    BitField("stream_engine.stream.stream.pingpong_last_index",             width=PORT_LAST_INDEX,      port="mse_pingpong_last_index",meaning="切换pingpong所需的last_index",remark=""),
                    BitField("",                                                            width=TRANSACTION_SPATIAL_SIZE_LOG_WIDTH,   port="mse_transaciton_layout_size_log"   ,  meaning="码流生成器计算: log2(idx2_size)和log2(idx1_size*idx2_size)"),
                    BitField("",                                                            width=TRANSACTION_TOTAL_SIZE_WIDTH,         port="mse_transaciton_total_size",          meaning="码流生成器计算: idx0_size*idx1_size*idx2_size)"),
                    
                    BitField("stream_engine.stream.memory_AG.idx",                      width=IDX_WIDTH,                                port="mem_inport_src_encode",       meaning="idx0、idx1、idx2来自于DRAM LC,配置点为“DRAM_LC.LCx”时对应码流为x(x的范围从0-7)" ,           remark="" ),
                    BitField("stream_engine.stream.memory_AG.idx_enable",               width=AG_PORT_NUM,                              port="mse_mem_idx_enable",meaning="DRAM AG的3个port使能情况，置为0的port的所有其他配置位都不生效",remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_keep_mode",            width=AG_PORT_NUM,                              port="mse_mem_idx_keep_mode",meaning="keep：1，buffer：0",remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_keep_last_index",      width=MEM_IDX_KEEP_LAST_INDEX_WIDTH,            port="mse_mem_idx_keep_last_index" ,meaning="DRAM AG的3个port各自所需的last index",remark="7"),
                    
                    BitField("stream_engine.stream.buffer_AG.spatial_stride",      width=BUF_STRIDE,           port="mse_buf_spatial_stride",      meaning="buffer ag展开的多路offset" ,remark=""),
                    BitField("stream_engine.stream.buffer_AG.spatial_size",        width=MSE_BUF_SIZE_WIDTH,   port="mse_buf_spatial_size",        meaning="buffer ag需要展开的地址个数", remark=""),
                    BitField("stream_engine.stream.buffer_AG.idx_enable",    width=BUFFER_AG_PORT_NUM,                              port="mse_buf_idx_enable",         meaning="buffer AG的2个port的使能情况", remark=""),
                    BitField("stream_engine.stream.buffer_AG.idx_keep_mode",    width=BUFFER_AG_PORT_NUM,                              port="mse_buf_idx_keep_mode",   meaning="keep：1，buffer：0", remark=""),
                    BitField("stream_engine.stream.buffer_AG.idx_keep_last_index",    width=BUF_IDX_KEEP_LAST_INDEX_WIDTH,            port="mse_buf_idx_keep_last_index",meaning="buffer AG的2个port各自所需的last index",remark="7"),
                    
                ],
            },
            {
                "title": "1*Write Memory Stream Engine",
                "total_bits": 1024,            
                "pack_direction": "lsb_to_msb",
                "fields": [
                    # 配置点
                    BitField("",                                                            width=1,                    port="mse_enable",           meaning="配置了就为1没配就为0",               remark=""),
                    BitField("stream_engine.stream.memory_AG.mode",                         width=1,                    meaning="read: 需要配padding和tailing  write: 不需要配",               remark=""),
                    BitField("stream_engine.stream.memory_AG.base_addr",                    width=ADDR_WIDTH,           port="mse_stream_base_addr",        meaning="基地址",                      remark="" ),
                    BitField("stream_engine.stream.memory_AG.idx_size",                     width=IDX_SIZE_WIDTH,       port="mse_transaciton_layout_size",meaning="idx0_size、idx1_size、idx2_size，实际值减1"  ),
                    BitField("stream_engine.stream.memory_AG.dim_stride",                   width=STRIDE_WIDTH,         port="mse_transaciton_mult",        meaning="idx0_stride、idx1_stride、idx2_stride"    ),
                    BitField("stream_engine.stream.memory_AG.address_remapping",            width=REMAP_WIDTH,          port="mse_map_matrix_b",            meaning="地址重映射表/掩码"     ,remark="[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]" ),
                    
                    BitField("stream_engine.stream.stream.ping_pong",                       width=PING_PONG_WIDTH,      port="mse_pingpong_enable",         meaning="ping-pong使能", remark=""),
                    BitField("stream_engine.stream.stream.pingpong_last_index",             width=PORT_LAST_INDEX,      port="mse_pingpong_last_index",meaning="切换pingpong所需的last_index",remark=""),
                    BitField("",                                                            width=TRANSACTION_SPATIAL_SIZE_LOG_WIDTH,   port="mse_transaciton_layout_size_log"   ,  meaning="码流生成器计算: log2(idx2_size)和log2(idx1_size*idx2_size)"),
                    BitField("",                                                            width=TRANSACTION_TOTAL_SIZE_WIDTH,         port="mse_transaciton_total_size",          meaning="码流生成器计算: idx0_size*idx1_size*idx2_size)"),
                    
                    BitField("stream_engine.stream.memory_AG.idx",                      width=IDX_WIDTH,                                port="mem_inport_src_encode",       meaning="idx0、idx1、idx2来自于DRAM LC,配置点为“DRAM_LC.LCx”时对应码流为x(x的范围从0-7)" ,           remark="" ),
                    BitField("stream_engine.stream.memory_AG.idx_enable",               width=AG_PORT_NUM,                              port="mse_mem_idx_enable",meaning="DRAM AG的3个port使能情况，置为0的port的所有其他配置位都不生效",remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_keep_mode",            width=AG_PORT_NUM,                              port="mse_mem_idx_keep_mode",meaning="keep：1，buffer：0",remark=""),
                    BitField("stream_engine.stream.memory_AG.idx_keep_last_index",      width=MEM_IDX_KEEP_LAST_INDEX_WIDTH,            port="mse_mem_idx_keep_last_index" ,meaning="DRAM AG的3个port各自所需的last index",remark="7"),
                    
                    BitField("stream_engine.stream.buffer_AG.spatial_stride",      width=BUF_STRIDE,           port="mse_buf_spatial_stride",      meaning="buffer ag展开的多路offset" ,remark=""),
                    BitField("stream_engine.stream.buffer_AG.spatial_size",        width=MSE_BUF_SIZE_WIDTH,   port="mse_buf_spatial_size",        meaning="buffer ag需要展开的地址个数", remark=""),
                    BitField("stream_engine.stream.buffer_AG.idx_enable",    width=BUFFER_AG_PORT_NUM,                              port="mse_buf_idx_enable",         meaning="buffer AG的2个port的使能情况", remark=""),
                    BitField("stream_engine.stream.buffer_AG.idx_keep_mode",    width=BUFFER_AG_PORT_NUM,                              port="mse_buf_idx_keep_mode",   meaning="keep：1，buffer：0", remark=""),
                    BitField("stream_engine.stream.buffer_AG.idx_keep_last_index",    width=BUF_IDX_KEEP_LAST_INDEX_WIDTH,            port="mse_buf_idx_keep_last_index",meaning="buffer AG的2个port各自所需的last index",remark="7"),
                    
                ],
            },
            {
                "title": "1*Neighbor Stream Engine",
                "total_bits": 1024,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("n2n.enable",              width=N2N_MODE,           port="nse_enable"           ,       remark=""),
                    BitField("n2n.mem_loop",            width=MEM_LOOP_WIDTH,       port="nse_cnt_size"         ,       remark=""),
                    BitField("n2n.src_slice_sel",       width=1,                    port="nse_in_src_slice_sel" ,       meaning="1表示跳4个slice，0表示不跳", remark=""),
                    BitField("n2n.dst_slice_sel",       width=1,                    port="nse_out_dst_slice_sel",       remark=""),
                    BitField("n2n.ping_pong",           width=1,                    port="nse_pingpong_enable",         remark=""),
                ],
            },
            {
                "title": "6*Buffer_Manager_Cluster",
                "total_bits": 1024,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("buffer_config.buffer.enable",             width=1,                            port="buffer_enable"        ,meaning=""),
                    BitField("buffer_config.buffer.dst_port",           width=1,                            port="buf_wr_src_id"        ,meaning="0: SpecArray, 1: GeneArray"),
                    BitField("buffer_config.buffer.buffer_life_time",   width=BUFFER_LIFE_TIME_WIDTH,       port="buffer_life_time"     ,meaning="buffer被复用多少次"),
                    BitField("buffer_config.buffer.mode",               width=1,                            port="buffer_mode"          ,meaning="mode0:for(life)for(row) mode1:for(row)for(life)"),
                    BitField("buffer_config.buffer.mask",               width=BUFFER_BANK_NUM,              port="buffer_mask"          ,meaning=""),
                ],
            }
        ]
    },
    {
        "group_title": "SA",
        "sections": [
            {
                "title": "1*PE",
                "total_bits": 1024, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("special_array.data_type", width=SA_PE_COMP_DATA_TYPE_WIDTH, port="sa_pe_computation_data_type"  ,meaning="fp16:1 int8:0"),
                    BitField("special_array.index_end", width=SA_PE_TRANSOUT_LAST_INDEX,  port="sa_pe_config_last_index"      ,meaning="PE的输出什么时候有效"),
                ],
            },
            {
                "title": "3*Inport",
                "total_bits": 1024, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("special_array.inport.enbale",                 width=1,                        port="sa_inport_nable",                 remark=""),
                    BitField("special_array.inport.pingpong_en",            width=1,                        port="sa_inport_pingpong_en",           meaning=""),
                    BitField("special_array.inport.pingpong_last_index",    width=3,                        port="sa_inport_pingpong_last_index",   meaning=""),
                ],
            },
            {
                "title": "1*Outport",
                "total_bits": 1024,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("special_array.outport.mode",          width=1,    port="sa_outport_major"            ,meaning="col：1  row:0"),
                    BitField("special_array.outport.fp32to16",      width=1,    port="sa_outport_fp32to16"         ,meaning="true: 1 false: 0"),
                ],
            },
        ]
    },
    {
        "group_title": "GA",
        "sections": [
            {
                "title": "16*PE",
                "total_bits": 1024, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("special_array.data_type", width=SA_PE_COMP_DATA_TYPE_WIDTH, port="sa_pe_computation_data_type"  ,meaning="fp16:1 int8:0"),
                    BitField("special_array.index_end", width=SA_PE_TRANSOUT_LAST_INDEX,  port="sa_pe_config_last_index"      ,meaning="PE的输出什么时候有效"),
                ],
            },
            {
                "title": "3*Inport",
                "total_bits": 1024, 
                "pack_direction": "lsb_to_msb", 
                "fields": [
                    BitField("general_array.inport.enable",                 width=1,                        port="ga_inport_enable",                 remark=""),
                    BitField("general_array.inport.src_id",                 width=1,                        port="ga_inport_src_id",                meaning="0：buffer 1：special_array"),
                    BitField("general_array.inport.pingpong",               width=3,                        port="ga_inport_pingpong_en",           meaning=""),
                    BitField("general_array.inport.pingpong_last_index",    width=3,                        port="ga_inport_pingpong_last_index",   meaning=""),
                    BitField("general_array.inport.fp16to32",               width=3,                        port="ga_inport_fp16to32",              meaning=""),
                    BitField("general_array.inport.int32tofp",              width=3,                        port="ga_inport_int32tofp",             meaning=""),
                ],
            },
            {
                "title": "1*Outport",
                "total_bits": 1024,
                "pack_direction": "lsb_to_msb",
                "fields": [
                    BitField("general_array.outport.enable",          width=1,    port="ga_outport_enable"            ,meaning=""),
                    BitField("general_array.outport.src_id",          width=1,    port="ga_outport_src_id"         ,meaning="选择outport接哪个PE"),
                    BitField("general_array.outport.fp16to32",          width=1,    port="ga_outport_fp32to16"            ,meaning=""),
                    BitField("general_array.outport.int32tofp",          width=1,    port="ga_outport_int32to8"         ,meaning=""),
                ],
            },
        ]
    }
]

if __name__ == "__main__":
    build_register_map_excel("register_map_with_groups.xlsx", groups)
    print("生成完成：register_map_with_groups.xlsx")