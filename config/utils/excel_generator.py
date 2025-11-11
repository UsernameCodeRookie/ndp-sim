import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from config.utils.config_parameters import AGMode, BAGMode

def _is_list_like(x):
    return isinstance(x, (list, tuple))

def _to_cell_value(v):
    try:
        return v.value  # Enum
    except AttributeError:
        return v

def multiple_dicts_to_excel(data_dict, output_dir="data", file_name="multiple_sheets.xlsx",
                            pad_with=0, max_col_width=80, padding=2):
    """
    - list/tuple 字段横向展开
    - 表头按“最大长度”展开并合并居中
    - Enum 写入其 .value
    - 自动列宽；对合并表头做二次校正，保证标题不被截断
    """
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, file_name)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in data_dict.items():
        if not rows:
            print(f"警告：{sheet_name} 无数据，跳过生成")
            continue

        ws = wb.create_sheet(title=sheet_name)

        # 1) 保持第一行键顺序
        headers = list(rows[0].keys())

        # 2) 统计每个字段的展开列数
        spans = {}
        for h in headers:
            mx = 1
            for r in rows:
                v = r.get(h, None)
                if _is_list_like(v):
                    mx = max(mx, len(v))
            spans[h] = mx

        # 3) 写“展开后的表头”占位行
        expanded_header_row = []
        header_start_col = {}
        col_ptr = 1
        for h in headers:
            span = spans[h]
            header_start_col[h] = col_ptr
            expanded_header_row.append(h)
            for _ in range(span - 1):
                expanded_header_row.append(None)
            col_ptr += span
        ws.append(expanded_header_row)

        # 4) 写数据行（对齐展开 & 右补齐）
        for r in rows:
            row_out = []
            for h in headers:
                span = spans[h]
                v = r.get(h, None)
                if _is_list_like(v):
                    vals = [_to_cell_value(x) for x in v]
                    if len(vals) < span:
                        vals = vals + [pad_with] * (span - len(vals))
                    row_out.extend(vals)
                else:
                    row_out.append(_to_cell_value(v))
                    if span > 1:
                        row_out.extend([pad_with] * (span - 1))
            ws.append(row_out)

        # 5) 合并表头 & 居中
        for h in headers:
            span = spans[h]
            start_col = header_start_col[h]
            if span > 1:
                ws.merge_cells(start_row=1, start_column=start_col,
                               end_row=1,   end_column=start_col + span - 1)
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')

        # 6) 自动列宽（逐列扫描最大字符串长度）
        #    先给每列一个初始宽度
        num_cols = ws.max_column
        col_widths = {}
        for c in range(1, num_cols + 1):
            max_len = 0
            for cell in ws.iter_cols(min_col=c, max_col=c, min_row=1, max_row=ws.max_row, values_only=True):
                for v in cell:
                    if v is None:
                        continue
                    s = str(v)
                    if "\n" in s:
                        s = max(s.splitlines(), key=len)
                    max_len = max(max_len, len(s))
            # 基础列宽 = 最大长度 + padding，限制最大
            col_widths[c] = min(max_col_width, max_len + padding)

        # 7) 针对合并表头做“总宽度校正”：保证合并后的总宽 >= 表头文本长度 + padding
        for h in headers:
            span = spans[h]
            start_col = header_start_col[h]
            header_len = len(str(h)) + padding
            if span == 1:
                # 单列：直接取 max(当前宽, header_len)
                col_widths[start_col] = max(col_widths[start_col], min(max_col_width, header_len))
            else:
                cols = list(range(start_col, start_col + span))
                current_total = sum(col_widths[c] for c in cols)
                if current_total < header_len:
                    deficit = header_len - current_total
                    # 平均摊到各列
                    add_each = deficit / span
                    for c in cols:
                        col_widths[c] = min(max_col_width, col_widths[c] + add_each)

        # 8) 应用列宽
        for c, w in col_widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(output_path)
    print(f"多工作表Excel已生成: {output_path}")
    return output_path


# 示例
if __name__ == "__main__":
    rd_mse_params_1 = {
        "mse_enable": 1,

        "mse_mem_idx_mode": [AGMode.KEEP, AGMode.CONSTANT, AGMode.BUFFER],
        "mse_mem_idx_keep_last_index": [2, 0, 4],
        "mem_inport_src_id": [0, 0, 8],
        "mse_mem_idx_constant": [0, 0, 0],

        "mse_buf_idx_mode": [BAGMode.KEEP, BAGMode.BUFFER],
        "mse_buf_idx_keep_last_index": [5, 0],

        "mse_pingpong_enable": 1,
        "mse_pingpong_last_index": 3,

        "mse_stream_base_addr": 0x20_0000,
        "mse_transaciton_layout_size": [1, 32, 4],
        "mse_transaciton_layout_size_log": [0, 0, 5],
        "mse_transaciton_total_size": 1*4*32,
        "mse_transaciton_mult": [4*64, 4, 4*64*9],  # [256, 4, 2304]
        "mse_map_matrix_b": list(range(16))[::-1],

        "mse_padding_reg_value": 0,
        "mse_padding_valid": [0, 0, 0],
        "mse_padding_low_bound": [0, 0, 0],
        "mse_padding_up_bound": [63, 31, 31],

        "mse_branch_valid": [0, 0, 0],
        "mse_branch_low_bound": [0, 0, 0],
        "mse_branch_up_bound": [63, 31, 31],

        "mse_buf_spatial_stride": list(range(16))[::-1],
        "mse_buf_spatial_size": 16 - 1,
    }
    all_data = {"RD_MSE": [rd_mse_params_1]}
    multiple_dicts_to_excel(all_data)